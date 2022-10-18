# -*- coding: utf-8 -*-
"""
Created on Mon Sep 12 14:02:45 2022

@author: ian88
"""

import os
import colour
import numpy as np
import pandas as pd
import openpyxl
from colour.utilities import as_float
from math import log10 as log10
from math import log as log

'''
'AUO_native'
XYZR = np.array([391.53878, 174.130524,2.491597]) #RGB 1023, 0, 0
XYZG = np.array([144.465195, 469.701599, 45.446548]) #RGB 0,1023, 0
XYZB = np.array([151.512985, 51.407036, 855.453796]) #RGB 0, 0, 1023
XYZW = np.array([682.830505, 689.485046, 903.17157]) #RGB 1023, 1023, 1023
XYZ0 = np.array([0.489888, 0.499837, 0.980175]) #RGB 0, 0, 0

'''    
'INN_native'
XYZR = np.array([302.417816, 138.637573, 4.292623]) #RGB 1023, 0, 0
XYZG = np.array([179.312927, 443.643799, 47.899467]) #RGB 0,1023, 0
XYZB = np.array([106.868942, 39.873371, 606.846863]) #RGB 0, 0, 1023
XYZW = np.array([594.82959, 625.800476, 678.746826]) #RGB 1023, 1023, 1023
XYZ0 = np.array([0.1892, 0.207, 0.3545]) #RGB 0, 0, 0


M_sRGB_RGB_XYZ = np.array(
            [[0.4124, 0.3576, 0.1805],
             [0.2126, 0.7152, 0.0722],
             [0.0193, 0.1192, 0.9505]])
    
M_ADOBE_RGB_XYZ = np.array(
            [[0.5767309, 0.1855540, 0.1881852],
             [0.2973769, 0.6273491, 0.0752741],
             [0.0270343, 0.0706872, 0.9911085]])
        
M_BT2020_RGB_XYZ = np.array(
            [[0.6370, 0.1446, 0.1689],
             [0.2627, 0.6780, 0.0593],
             [0.0, 0.0281, 1.0610]])

gamut_matrix = {'sRGB': M_sRGB_RGB_XYZ, 'ADOBE': M_ADOBE_RGB_XYZ, 'BT2020': M_BT2020_RGB_XYZ}

def IS_RGB(R,G,B):
    if(R+G+B == np.max([R,G,B])):
        return 1   
    
def write_color_data(data, r,c, mode, exs = None):
    if (mode == 1):
        for d in range(0, len(data)):
            for i in range(len(data[d])):
                exs.cell(r+d+2, i+c).value=np.round(data[d][i], 4)
    elif (mode == 2):
        for d in range(0,len(data)):
            exs.cell(d+2, c).value=np.round(data[d], 4)

def Build_Matrix_RGB_to_XYZ(XYZ_R, XYZ_G, XYZ_B, XYZ_0, Y_max):
    XYZR_nor = (XYZ_R-XYZ_0)/Y_max
    XYZG_nor = (XYZ_G-XYZ_0)/Y_max
    XYZB_nor = (XYZ_B-XYZ_0)/Y_max
    
    XYZR_nor_sum = sum(XYZR_nor)
    XYZG_nor_sum = sum(XYZG_nor)
    XYZB_nor_sum = sum(XYZB_nor)

    xR = XYZR_nor[0]/XYZR_nor_sum
    xG = XYZG_nor[0]/XYZG_nor_sum
    xB = XYZB_nor[0]/XYZB_nor_sum
    
    yR = XYZR_nor[1]/XYZR_nor_sum
    yG = XYZG_nor[1]/XYZG_nor_sum
    yB = XYZB_nor[1]/XYZB_nor_sum
            
    zR = XYZR_nor[2]/XYZR_nor_sum
    zG = XYZG_nor[2]/XYZG_nor_sum
    zB = XYZB_nor[2]/XYZB_nor_sum

    Xi = np.array(
        [[xR, xG, xB],
         [yR, yG, yB],
         [zR, zG, zB]])        
        
    Sm = np.array(
        [[XYZR_nor_sum, 0, 0],
         [0, XYZG_nor_sum, 0],
         [0, 0, XYZB_nor_sum]])
    
    M = Xi.dot(Sm)
    return M

def chromatic_adaptation(Y_max, XYZ, XYZ_ws, XYZ_wr):#color adaptation using Von Kries
        XYZ_nor = XYZ/Y_max
        L_A1 = Y_max/np.pi
        L_A2 = Y_max/np.pi
        
        XYZ_cp = colour.chromatic_adaptation(
                XYZ_nor, XYZ_ws, XYZ_wr, method = 'CMCCAT2000', L_A1 = L_A1, L_A2 = L_A2)
            
        return XYZ_cp*Y_max

def CCT_to_xyz_nor(CCT):
     xy = np.array(colour.CCT_to_xy(CCT))
     xyz = np.append(xy, 1.0-xy[0]-xy[1])
     xyz*= 1.0/xy[1]
     return xyz
 
def W_Transform(matrix_native_RGB_XYZ, CCT_target):
    """
        matrix_native_RGB_XYZ: matrix of native RGB to XYZ\n
        CCT_target: ex 5000、6500、7000
    """
    target_xyz = CCT_to_xyz_nor(CCT_target)
    M1 = np.linalg.inv(matrix_native_RGB_XYZ)
    K = M1.dot(np.array(target_xyz).T)
    return K
 
def XYZ_STD(gamut_matrix, trc, cct, Y_max, dicom = False, dicom_gamma = []):
    XYZ_l = []
    XYZ_RGB_l = []
    XYZ_W_l = []
    
    w_k = W_Transform(gamut_matrix, cct)
    matrix_cct =  np.array(
        [[w_k[0], 0, 0],
         [0,w_k[1], 0],
         [0, 0, w_k[2]]])
    gamut_matrix = gamut_matrix.dot(matrix_cct)
    
    for R in range(0, 1056, 128):
        for G in range(0, 1056, 128):
            for B in range(0, 1056, 128):
                if(R == 1024):
                    r = 1023
                else:            
                    r = R
                if(G == 1024):
                    g = 1023
                else:            
                    g = G
                if(B == 1024):
                    b = 1023
                else:            
                    b = B  
                if(dicom == True):
                    XYZ = RGB_to_XYZ_dicom(gamut_matrix, [r,g,b], dicom_gamma, Y_max)
                else:  
                    XYZ = RGB_to_XYZ(gamut_matrix, [r,g,b], trc, Y_max)

                if(IS_RGB(r,g,b)):
                    XYZ_RGB_l.append(XYZ)    
                elif(r==g==b):                   
                    continue
                else:
                    XYZ_l.append(XYZ)  
                    
    XYZ_l = np.concatenate([XYZ_l, XYZ_RGB_l],axis=0)
    
    for W in range(0, 1056, 64): 
        if(W == 1024):
            w = 1023
        else:
            w = W
            
        if(dicom == True):
            XYZ = RGB_to_XYZ_dicom(gamut_matrix, [w,w,w], dicom_gamma, Y_max)
        else:  
            XYZ = RGB_to_XYZ(gamut_matrix, [w,w,w], trc, Y_max)
        XYZ_W_l.append(XYZ)  
     
    XYZ_l = np.concatenate([XYZ_l, XYZ_W_l],axis=0)
    return XYZ_l
    
def RGB_to_XYZ(Matrix, RGB, trc, Y_max):
    rgb = np.array(RGB)/1023
    v = [rgb[0]**trc, 
         rgb[1]**trc,
         rgb[2]**trc] 
    
    XYZ = Matrix.dot(v)*Y_max
    return XYZ 

def RGB_to_XYZ_dicom(Matrix, RGB, dicom_gamma, Y_max):
    rgb = np.array(RGB)
    v = [1023.0*dicom_gamma[rgb[0]], 
         1023.0*dicom_gamma[rgb[1]],
         1023.0*dicom_gamma[rgb[2]]] 
    
    XYZ = Matrix.dot(v)*Y_max
    return XYZ 

def f(t):
    t_o = as_float(np.where(
            t > 0.008856,
            t ** (1.0 / 3.0),
            (903.3 * t + 16.0) / 116.0,
    ))
    return t_o

def XYZ_to_Lab(XYZ, XYZ_r):
    xr = XYZ[0]/XYZ_r[0]
    yr = XYZ[1]/XYZ_r[1]
    zr = XYZ[2]/XYZ_r[2]
    
    fx = f(xr)
    fy = f(yr)
    fz = f(zr)
    
    L = 116.0 * fy - 16.0
    a = 500.0 * (fx - fy)
    b = 200.0 * (fy - fz)
    
    return np.array([L, a, b]) 

def xy_to_XYZ(xy, Y):
    X = Y*xy[0]/xy[1]
    Z = Y*(1.0-xy[0]-xy[1])/xy[1]
    
    return np.array([X, Y, Z]) 

def L_to_JND(L):
    JND = 0
    JND = 71.498068 + 94.593053 * log10(L) + 41.912053 * (pow(log10(L), 2)) + 9.8247004 * (pow(log10(L), 3))+ 0.28175407 * (pow(log10(L), 4)) - 1.1878455 * (pow(log10(L), 5))- 0.18014349 * (pow(log10(L), 6)) + 0.14710899 * (pow(log10(L), 7)) - 0.017046845 * (pow(log10(L), 8));
    return JND

def JND_to_L(JND):
    j = JND
    d = -1.3011877 + 0.080242636 * log(j) + 0.13646699 * (pow(log(j), 2)) - 0.025468404 * (pow(log(j), 3)) + 0.0013635334 * (pow(log(j), 4));
    m = 1 - 0.025840191 * log(j) - 0.10320229 * (pow(log(j), 2)) + 0.028745620 * (pow(log(j), 3)) - 0.0031978977 * (pow(log(j), 4)) + 0.00012992634 * (pow(log(j), 5));
    log10L = d / m
    L = pow(10, log10L)
    return L

def j_p(Jmax, Jmin, n, p):
    jp = Jmin + (p / (pow(2, n) - 1)) * (Jmax - Jmin)
    return jp;

def dicom_XYZ(max_Lv):
    target_dicom_list = []
    J_max = L_to_JND(max_Lv)
    J_min = L_to_JND(0.01)
    n = 10;

    for p in range(0, 1024) :
        jp = j_p(J_max, J_min, n, p);
        if (jp <= 0):
            L_D = 0
        else:
            L_D = JND_to_L(jp)
            L_D /= max_Lv
        
        target_dicom_list.append(L_D);
	
    target_dicom_list[0] = 0;
    target_dicom_list[len(target_dicom_list) - 1] = 1;
    return target_dicom_list;


def main():
    Matrix_Native = Build_Matrix_RGB_to_XYZ(XYZR, XYZG, XYZB, XYZ0, XYZW[1])

    ex_file = './a.xlsx'
    ex = openpyxl.load_workbook(ex_file)
    
    gamut_list = ['Native', 'sRGB', 'ADOBE', 'BT2020']
    for g in  gamut_list:
        ca_path = '../color/' + g + '/'
        for fs in os.listdir(ca_path):
            cct = int(fs[0:4])
            trc = fs[5:9]
            
            if trc == 'dico':
                ca_file = ca_path +str(cct) + '_' + 'dicom.csv'
            else:
                ca_file = ca_path +str(cct) + '_' + str(trc) + '.csv'

            print(ca_file)
            try:  
                df = pd.read_csv(ca_file)
                sourcedata = df.values
                XYZ_ca_L = np.vstack([sourcedata[:, 3:6]])
                Y_max = XYZ_ca_L[-1][1]                
            except:
                print('Invalid Calibrated Data!')
                os.system("pause")     
            
            if(g=='Native'):
                if(trc =='dico'):
                    dicom_gamma = dicom_XYZ(Y_max)
                    XYZ_std_L = XYZ_STD(Matrix_Native, 0, cct, Y_max, True, dicom_gamma)                  
                else:
                    XYZ_std_L = XYZ_STD(Matrix_Native, float(trc), cct, Y_max)
  
            else:
                if(trc =='dico'):
                    dicom_gamma = dicom_XYZ(Y_max)
                    XYZ_std_L = XYZ_STD(gamut_matrix[g], 0, cct, Y_max, True, dicom_gamma)        
                else:
                    XYZ_std_L = XYZ_STD(gamut_matrix[g], float(trc), cct, Y_max)
               
            ref_white = XYZ_std_L[-1]
            delta_E_L = []
            
            r = XYZ_ca_L[-1][1] / XYZ_std_L[-1][1] 
            XYZ_std_L*=r
            
            for i in range(0, len(XYZ_std_L)):
                XYZ_ca = XYZ_ca_L[i]
                XYZ_std = XYZ_std_L[i]
                Lab_ca = XYZ_to_Lab(XYZ_ca, ref_white)
                Lab_std = XYZ_to_Lab(XYZ_std, ref_white)
                delta_E_L.append(colour.delta_E_CIE2000(Lab_ca, Lab_std))
            
            exs = ex.get_sheet_by_name("data") 
            if(trc =='dico'):
                exs.cell(2, 1).value = 'dicom'
                file = './' + g + "_" + str(cct) + '_dicom.xlsx'
            else:
                exs.cell(2, 1).value = float(trc)
                file = './' + g + "_" + str(cct) + '_' + str(trc) + '.xlsx'
            write_color_data(XYZ_std_L, 0, 5, 1, exs)
            write_color_data(XYZ_ca_L, 0, 25, 1, exs)
            write_color_data(delta_E_L, 0, 31, 2, exs)
            ex.save(file)
            
    print("\nFinish!")
    os.system("pause")
main()