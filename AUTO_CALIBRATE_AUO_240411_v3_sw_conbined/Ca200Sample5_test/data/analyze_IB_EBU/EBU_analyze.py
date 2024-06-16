# -*- coding: utf-8 -*-
"""
Created on Mon Apr 10 14:04:06 2023

@author: ian88
"""

import os
import colour
import numpy as np
import pandas as pd
import openpyxl
from colour.utilities import as_float
from math import log10 as log10
from math import log as log
from math import sqrt
import math

'AUO_native'
XYZR = np.array([401.187317, 178.678558, 3.199178]) #RGB 1023, 0, 0
XYZG = np.array([151.480087, 519.197937, 51.873894]) #RGB 0,1023, 0
XYZB = np.array([166.068314, 54.240467, 891.829224]) #RGB 0, 0, 1023
XYZW = np.array([718.80365, 750.381531, 947.157104]) #RGB 1023, 1023, 1023
XYZ0 = np.array([0.505993, 0.542448, 1.010076]) #RGB 0, 0, 0

M_sRGB_RGB_XYZ = np.array(
            [[0.4124, 0.3576, 0.1805],
             [0.2126, 0.7152, 0.0722],
             [0.0193, 0.1192, 0.9505]])
    
M_ADOBE_RGB_XYZ = np.array(
            [[0.5767309, 0.1855540, 0.1881852],
             [0.2973769, 0.6273491, 0.0752741],
             [0.0270343, 0.0706872, 0.9911085]])
        
M_BT2020_RGB_XYZ = np.array(
            [[0.6370, 0.1446, 0.1689],
             [0.2627, 0.6780, 0.0593],
             [0.0, 0.0281, 1.0610]])

gamut_matrix_list = {'sRGB': M_sRGB_RGB_XYZ, 'ADOBE': M_ADOBE_RGB_XYZ, 'BT2020': M_BT2020_RGB_XYZ}

Yuv_list = np.array(
    [[9.60, 0.2526, 0.5013], [37.7, 0.2365, 0.4931], [29.8, 0.2365, 0.4850],
     [29.9, 0.1805, 0.5453], [29.8, 0.1630, 0.4552], [30.1, 0.2088, 0.4155],
     [13.4, 0.1816, 0.5205], [19.3, 0.3246, 0.4972], [43.6, 0.1504, 0.5329],
     [17.2, 0.1789, 0.3707], [6.50, 0.3047, 0.4898], [20.0, 0.1461, 0.5320],
     [6.00, 0.1828, 0.3420], [43.4, 0.2724, 0.5274], [20.0, 0.2348, 0.4031]])

def write_color_data(data, r,c, mode, exs = None):
    if (mode == 1):
        for d in range(0, len(data)):
            for i in range(len(data[d])):
                exs.cell(r+d+2, i+c).value=np.round(data[d][i], 4)
    elif (mode == 2):
        for d in range(0,len(data)):
            exs.cell(d+2, c).value=np.round(data[d], 4)

def CCT_to_xyz_nor(CCT):
     xy = np.array(colour.CCT_to_xy(CCT))
     xyz = np.append(xy, 1.0-xy[0]-xy[1])
     xyz*= 1.0/xy[1]
     return xyz
 
def W_Transform(matrix_native_RGB_XYZ, CCT_target):
    """
        matrix_native_RGB_XYZ: matrix of native RGB to XYZ\n
        CCT_target: ex 5000、6500、7000
    """
    target_xyz = CCT_to_xyz_nor(CCT_target)
    M1 = np.linalg.inv(matrix_native_RGB_XYZ)
    K = M1.dot(np.array(target_xyz).T)
    return K

def RGB_to_XYZ(Matrix, RGB, trc, Y_max):
    rgb = np.array(RGB)/1023.0
    v = [rgb[0]**trc, 
         rgb[1]**trc,
         rgb[2]**trc] 
    
    XYZ = Matrix.dot(v)*Y_max
    return XYZ 

def XYZ_to_RGB(Matrix, XYZ, trc, Y_max):
    XYZ = np.asarray(XYZ)
    XYZ /= Y_max
    m =  np.linalg.inv(Matrix)
    rgb = m.dot(XYZ)
    
    for i in range(0, len(rgb)):
        if rgb[i] < 0:
            rgb[i] = 0
        
    R = np.round(pow((rgb[0]), 1/trc) * 1023, 0)
    G = np.round(pow((rgb[1]), 1/trc) * 1023, 0)
    B = np.round(pow((rgb[2]), 1/trc) * 1023, 0)
    
    return [R, G, B]

def Generate_Target_XYZ(matrix_native_RGB_XYZ, RGB_list, cct, trc, Y_max):
    XYZ_L = []
    w_k= W_Transform(matrix_native_RGB_XYZ, cct)
    matrix_cct = np.array(
        [[w_k[0], 0, 0],
         [0,w_k[1], 0],
         [0, 0, w_k[2]]])
    matrix = matrix_native_RGB_XYZ.dot(matrix_cct)
     
    for i in range(0,len(RGB_list)):
        XYZ = RGB_to_XYZ(matrix, RGB_list[i], trc, Y_max)
        XYZ_L.append(XYZ)
    
    return XYZ_L

def Yuv_to_XYZ(Yuv, Y_max):
    x = 9 * Yuv[1] / (6 * Yuv[1] - 16 * Yuv[2] + 12)
    y = 4 * Yuv[2] / (6 * Yuv[1] - 16 * Yuv[2] + 12)
    #xy = colour.Luv_uv_to_xy([Yuv[1],Yuv[2]])
    #x = xy[0]
    #y = xy[1]
    Y = Y_max * Yuv[0] / 100.0
    X = Y * x / y
    Z = Y * (1 - x - y )/ y
    XYZ_list = [X, Y, Z]
     
    return XYZ_list

def f(t):
    t_o = as_float(np.where(
            t > 0.008856,
            t ** (1.0 / 3.0),
            (903.3 * t + 16.0) / 116.0,
    ))
    return t_o

def XYZ_to_Lab(XYZ, XYZ_r):
    xr = XYZ[0]/XYZ_r[0]
    yr = XYZ[1]/XYZ_r[1]
    zr = XYZ[2]/XYZ_r[2]
    
    fx = f(xr)
    fy = f(yr)
    fz = f(zr)
    
    L = 116.0 * fy - 16.0
    a = 500.0 * (fx - fy)
    b = 200.0 * (fy - fz)
    
    return np.array([L, a, b]) 

def XYZ_to_Luv(XYZ, XYZ_r):
    X = XYZ[0]
    Y = XYZ[1]
    Z = XYZ[2]
    
    X_r = XYZ_r[0]
    Y_r = XYZ_r[1]
    Z_r = XYZ_r[2]
    
    yr = XYZ[1]/XYZ_r[1]
    
    if X == 0:
        u_ = 0
    else:
        u_ = (4 * X) / (X + 15 * Y + 3 * Z)
    
    if X == 0:
        v_ = 0
    else:
        v_ = (9 * Y) / (X + 15 * Y + 3 * Z)
    
    if X_r == 0:
        u_r = 0
    else:
        u_r = (4 * X_r) / (X_r + 15 * Y_r + 3 * Z_r)
    
    if Y_r == 0:
        v_r = 0
    else:
        v_r = (9 * Y_r) / (X_r + 15 * Y_r + 3 * Z_r)
    
    
    L = np.where(yr > 0.008856, 116.0 * (math.pow(yr,(1.0 / 3.0))) - 16, 903.3 * yr)
    u = 13 * L *(u_ - u_r)
    v = 13 * L *(v_ - v_r)
    
    return np.array([L, u, v]) 

def delta_uv(Luv, Luv_r):
    d_uv = sqrt(math.pow(Luv[1] - Luv_r[1], 2) + math.pow(Luv[2] - Luv_r[2], 2))
    return d_uv

def delta_Euv(Luv, Luv_r):
    d_Euv = sqrt(math.pow(Luv[0] - Luv_r[0], 2) +  math.pow(Luv[1] - Luv_r[1], 2) + math.pow(Luv[2] - Luv_r[2], 2))
    return d_Euv 
    
def ADD_EBU(gamut, LUT, Yuv_list, trc, Y_max):
    LUT_new = LUT
    gamut_matrix = gamut_matrix_list[gamut]
    
    for i in range(0, len(Yuv_list)):
        Yuv = Yuv_list[i]
        XYZ = Yuv_to_XYZ(Yuv, Y_max)
        RGB = XYZ_to_RGB(gamut_matrix, XYZ, trc, Y_max)
        LUT_new.append(RGB)
        
    return LUT_new

def L_to_JND(L):
    JND = 0
    JND = 71.498068 + 94.593053 * log10(L) + 41.912053 * (pow(log10(L), 2)) + 9.8247004 * (pow(log10(L), 3))+ 0.28175407 * (pow(log10(L), 4)) - 1.1878455 * (pow(log10(L), 5))- 0.18014349 * (pow(log10(L), 6)) + 0.14710899 * (pow(log10(L), 7)) - 0.017046845 * (pow(log10(L), 8));
    return JND

def JND_to_L(JND):
    j = JND
    d = -1.3011877 + 0.080242636 * log(j) + 0.13646699 * (pow(log(j), 2)) - 0.025468404 * (pow(log(j), 3)) + 0.0013635334 * (pow(log(j), 4));
    m = 1 - 0.025840191 * log(j) - 0.10320229 * (pow(log(j), 2)) + 0.028745620 * (pow(log(j), 3)) - 0.0031978977 * (pow(log(j), 4)) + 0.00012992634 * (pow(log(j), 5));
    log10L = d / m
    L = pow(10, log10L)
    return L

def j_p(Jmax, Jmin, n, p):
    jp = Jmin + (p / (pow(2, n) - 1)) * (Jmax - Jmin)
    return jp;

def dicom_XYZ(min_Lv, max_Lv, RGB_list):
    target_dicom_list = []
    J_max = L_to_JND(max_Lv)
    J_min = L_to_JND(min_Lv)
    if J_min < 0:
        J_min = 0
        
    n = 10;
    
    xy = np.array(colour.CCT_to_xy(6500))
    ref_white = max_Lv*xy[0]/xy[1],  max_Lv,  max_Lv*(1-xy[0]-xy[1])/xy[1]

    for i in range(0, len(RGB_list)) :
        p = RGB_list[i][1]
        jp = j_p(J_max, J_min, n, p);
        if (jp <= 0):
            L_D = 0
        else:
            L_D = JND_to_L(jp)
            L_D /= max_Lv
        
        XYZ = []
        for j in range(0, 3):
            XYZ.append(ref_white[j] * L_D)
        target_dicom_list.append(XYZ);
	
    #target_dicom_list[0] = 0;
    #target_dicom_list[len(target_dicom_list) - 1] = 1;
    return target_dicom_list;

def main():
    mode = input("1 for genearate lut, 2 for analyze: ")
    
    LUT = []
    
    LUT.append([1023, 0, 0])
    LUT.append([0, 1023, 0])
    LUT.append([0, 0, 1023])
    LUT.append([1023, 1023, 1023])
    LUT.append([0, 1023, 1023])
    LUT.append([1023, 1023, 0])
    LUT.append([1023, 0, 1023])
    
    LUT.append([0, 0, 0])
    LUT.append([102, 102, 102])
    LUT.append([205, 205, 205])
    LUT.append([307, 307, 307])
    LUT.append([409, 409, 409])
    LUT.append([512, 512, 512])
    LUT.append([614, 614, 614])
    LUT.append([716, 716, 716])
    LUT.append([818, 818, 818])
    LUT.append([921, 921, 921])
    LUT.append([1023, 1023, 1023])
        
    LUT = ADD_EBU("sRGB", LUT, Yuv_list, 2.2, 100)
    
    LUT.append([0, 0, 0])
    LUT.append([102, 102, 102])
    LUT.append([205, 205, 205])
    LUT.append([307, 307, 307])
    LUT.append([409, 409, 409])
    LUT.append([512, 512, 512])
    LUT.append([614, 614, 614])
    LUT.append([716, 716, 716])
    LUT.append([818, 818, 818])
    LUT.append([921, 921, 921])
    LUT.append([1023, 1023, 1023])
    
    if mode == "1":
        XYZ = Generate_Target_XYZ(gamut_matrix_list["sRGB"], LUT, 6500, 2.2, 100)
        ex_file = './LUT.xlsx'
        ex = openpyxl.load_workbook(ex_file)
        exs = ex.get_sheet_by_name("Data")
        write_color_data(LUT, 0, 2, 1, exs)
        write_color_data(XYZ, 0, 6, 1, exs)
    
    else:
        ex_file = './report.xlsx'
        ex = openpyxl.load_workbook(ex_file)
        try:  
            ca_file = "./data.csv"
            df = pd.read_csv(ca_file)
            sourcedata = df.values
            XYZ_raw_L = np.vstack([sourcedata[:, 4:7]])
            XYZ_ca_L = np.vstack([sourcedata[:, 15:18]])
            CCT_raw_L = sourcedata[:, 11]
            CCT_ca_L = sourcedata[:, 22]
            Y_max = XYZ_ca_L[-1][1]                
        except:
            print('Invalid Raw Data!')
            os.system("pause")   
            
        XYZ_gamma = Generate_Target_XYZ(gamut_matrix_list["sRGB"], LUT[:33], 6500, 2.2, Y_max)
        XYZ_dicom = dicom_XYZ(XYZ_ca_L[33][1], XYZ_ca_L[-1][1], LUT[33:])
        XYZ_target_L = np.concatenate((XYZ_gamma,XYZ_dicom), axis=0)
        
        ref_white =  XYZ_target_L[-1]

        delta_uv_raw_L = []
        delta_uv_ca_L = []
        delta_E_raw_L = []
        delta_E_ca_L = []
        delta_E_2000_raw_L = []
        delta_E_2000_ca_L = []
        
        for i in range(0, len(XYZ_target_L)):
            XYZ_std = XYZ_target_L[i]
            XYZ_raw = XYZ_raw_L[i]
            XYZ_ca = XYZ_ca_L[i]
            
            Luv_std = XYZ_to_Luv(XYZ_std, ref_white)
            Luv_raw = XYZ_to_Luv(XYZ_raw, ref_white)
            Luv_ca = XYZ_to_Luv(XYZ_ca, ref_white)
            
            Lab_std = XYZ_to_Lab(XYZ_std, ref_white)
            Lab_raw = XYZ_to_Lab(XYZ_raw, ref_white)
            Lab_ca = XYZ_to_Lab(XYZ_ca, ref_white)
            
            delta_uv_raw_L.append(delta_uv(Luv_raw, Luv_std))
            delta_uv_ca_L.append(delta_uv(Luv_ca, Luv_std))
            delta_E_raw_L.append(delta_Euv(Luv_raw, Luv_std))
            delta_E_ca_L.append(delta_Euv(Luv_ca, Luv_std))
            delta_E_2000_raw_L.append(colour.delta_E_CIE2000(Lab_raw, Lab_std))
            delta_E_2000_ca_L.append(colour.delta_E_CIE2000(Lab_ca, Lab_std))
        
        exs = ex.get_sheet_by_name("Data")
        write_color_data(LUT, 0, 1, 1, exs)
        write_color_data(XYZ_target_L, 0, 5, 1, exs)
        write_color_data(XYZ_raw_L, 0, 15, 1, exs)
        write_color_data(XYZ_ca_L, 0, 282, 1, exs)
        write_color_data(CCT_raw_L, 0, 22, 2, exs)
        write_color_data(CCT_ca_L, 0, 35, 2, exs)
        write_color_data(delta_uv_raw_L, 0, 23, 2, exs)
        write_color_data(delta_E_raw_L, 0, 24, 2, exs)
        write_color_data(delta_E_2000_raw_L, 0, 25, 2, exs)
        write_color_data(delta_uv_ca_L, 0, 36, 2, exs)
        write_color_data(delta_E_ca_L, 0, 37, 2, exs)
        write_color_data(delta_E_2000_ca_L, 0, 38, 2, exs)
        
        
    ex.save(ex_file)
    
main()
