# -*- coding: utf-8 -*-
"""
Created on Mon Sep 12 14:02:45 2022

@author: Ian Chen
"""

import os
import colour
import numpy as np
import pandas as pd
import openpyxl
from colour.utilities import as_float
from math import log10 as log10
from math import log
from math import sqrt, exp
import math

'''
File structure
    - calibration_report.py
    - raw_color.csv
    - raw_color_lut.csv
    - color datas(csv)
    - color data_lut.csv
    - report folder
        - Native
        - sRGB
        - ADOBE
        - BT2020
        - overview.xlsx  
        
Data file format: %s_%d_%%s % gamut, cct , tone curve
ex: Native_6500_2.2, sRGB_9300_PQ, ADOBE_5000_DICOM

Supported color :
    gamut - Native, sRGB, ADOBE, BT2020
    cct - 3500, 4000, 5000, 5800, 6500,7500, 9300
    tone curve - gamma, DICOM, PQ, HLG
'''

RGB_max = 1023.0 #8-bit:255, 10-bit:1023

M_sRGB_RGB_XYZ = np.array(
            [[0.4124, 0.3576, 0.1805],
             [0.2126, 0.7152, 0.0722],
             [0.0193, 0.1192, 0.9505]])
    
M_ADOBE_RGB_XYZ = np.array(
            [[0.5767309, 0.1855540, 0.1881852],
             [0.2973769, 0.6273491, 0.0752741],
             [0.0270343, 0.0706872, 0.9911085]])
        
M_BT2020_RGB_XYZ = np.array(
            [[0.6370, 0.1446, 0.1689],
             [0.2627, 0.6780, 0.0593],
             [0.0, 0.0281, 1.0610]])

gamut_matrix = {'sRGB': M_sRGB_RGB_XYZ, 'ADOBE': M_ADOBE_RGB_XYZ, 'BT2020': M_BT2020_RGB_XYZ}
    

M_Bradford = np.array(
            [[0.8951000, 0.2664000, -0.1614000],
             [-0.7502000, 1.7135000, 0.0367000],
             [0.0389000, -0.0685000, 1.0296000]])

M_Bradford_inv = np.array(
            [[0.9869929, -0.1470543, 0.1599627],
             [0.4323053, 0.5183603, 0.0492912],
             [-0.0085287, 0.0400428, 0.9684867]])


def read_csv_data(file):
    try:  
        df = pd.read_csv(file)
        sourcedata = df.values
    except:
        print('Invalid Calibrated Data!')
        os.system("pause")  
    
    return sourcedata

def write_color_data(data, r,c, mode, exs = None):
    if (mode == 1):
        for d in range(0, len(data)):
            for i in range(len(data[d])):
                exs.cell(r+d+2, i+c).value=np.round(data[d][i], 4)
    elif (mode == 2):
        for d in range(0,len(data)):
            exs.cell(d+2, c).value=np.round(data[d], 4)


def Build_Matrix_RGB_to_XYZ(XYZ_R, XYZ_G, XYZ_B, XYZ_0, XYZ_W):
    xR = XYZ_R[0] / XYZ_R[1]
    yR = XYZ_R[1] / XYZ_R[1]
    zR = XYZ_R[2] / XYZ_R[1]
    
    xG = XYZ_G[0] / XYZ_G[1]
    yG = XYZ_G[1] / XYZ_G[1]
    zG = XYZ_G[2] / XYZ_G[1]
    
    xB = XYZ_B[0] / XYZ_B[1]
    yB = XYZ_B[1] / XYZ_B[1]
    zB = XYZ_B[2] / XYZ_B[1]
    
    Xi = np.array(
        [[xR, xG, xB],
         [yR, yG, yB],
         [zR, zG, zB]])        
        
    Sm = np.linalg.inv(Xi).dot(XYZ_W)
    
    
    M = np.array(
        [[xR * Sm[0], xG* Sm[1], xB* Sm[2]],
         [yR * Sm[0], yG* Sm[1], yB* Sm[2]],
         [zR * Sm[0], zG* Sm[1], zB* Sm[2]]])  
    
    return M


def Chromatic_Adaptation(XYZ, XYZ_W_S, XYZ_W_D):
    XYZ_W_S = np.array(XYZ_W_S)
    XYZ_W_D = np.array(XYZ_W_D)

    XYZ_W_S /= max(XYZ_W_S)
    XYZ_W_D /= max(XYZ_W_D)    
    cone_res_d = M_Bradford.dot(XYZ_W_D)
    cone_res_s = M_Bradford.dot(XYZ_W_S)
    
    m =  np.array(
        [[cone_res_d[0]/cone_res_s[0], 0, 0],
         [0, cone_res_d[1]/cone_res_s[1], 0],
         [0, 0, cone_res_d[2]/cone_res_s[2]]])   
     
    bradford_m =(M_Bradford_inv.dot(m)).dot(M_Bradford)
    XYZ_D = bradford_m.dot(XYZ)

    return XYZ_D, bradford_m


def CCT_to_xyz_nor(cct):
     xy = np.array(colour.CCT_to_xy(cct))
     xyz = np.append(xy, 1.0 - xy[0] - xy[1])
     
     xyz *= 1.0 / xy[1]
    
     return xyz
 
    
def W_Transform(matrix_native_RGB_XYZ, cct_target):
    """
        matrix_native_RGB_XYZ: matrix of native RGB to XYZ\n
        CCT_target: ex 5000、6500、7000
    """
    cct_target = int(cct_target)
    target_xyz = CCT_to_xyz_nor(cct_target)
    M1 = np.linalg.inv(matrix_native_RGB_XYZ)
    K = M1.dot(np.array(target_xyz).T)

    return K / max(K)
 

def Build_Matrix_Gamut_CCT_RGB_to_XYZ(matrix_gamut_target, cct):
    XYZ_R =  RGB_to_XYZ(matrix_gamut_target, [1023, 0, 0], 1, 0, 100)
    XYZ_G =  RGB_to_XYZ(matrix_gamut_target, [0, 1023, 0], 1, 0, 100)
    XYZ_B =  RGB_to_XYZ(matrix_gamut_target, [0, 0, 1023], 1, 0, 100)
    XYZ_W =  RGB_to_XYZ(matrix_gamut_target, [1023, 1023, 1023], 1, 0, 100)
    XYZ_0 =  RGB_to_XYZ(matrix_gamut_target, [0, 0, 0], 1, 0, 100)
    
    target_xyz = CCT_to_xyz_nor(cct)*100
    XYZ_R_b, bradford_m = Chromatic_Adaptation(XYZ_R, XYZ_W, target_xyz)
    XYZ_G_b, bradford_m = Chromatic_Adaptation(XYZ_G, XYZ_W, target_xyz)
    XYZ_B_b, bradford_m = Chromatic_Adaptation(XYZ_B, XYZ_W, target_xyz)
    XYZ_W_b, bradford_m = Chromatic_Adaptation(XYZ_W, XYZ_W, target_xyz)
    
    matrix_Gamut_CCT = Build_Matrix_RGB_to_XYZ(XYZ_R_b, XYZ_G_b, XYZ_B_b, XYZ_0, XYZ_W_b)

    return matrix_Gamut_CCT
    

def XYZ_STD(matrix_native_RGB_XYZ, gamut, cct, tone_curve, XYZ_0, XYZ_W, RGB_list):
    XYZ_list = []
    
    if gamut == "Native":
        matrix_gamut_target = matrix_native_RGB_XYZ
    else:
        matrix_gamut_target = gamut_matrix[gamut]
    
    color_matrix = Build_Matrix_Gamut_CCT_RGB_to_XYZ(matrix_gamut_target, cct)

    for i in range(0, len(RGB_list)):        
        XYZ_list.append(RGB_to_XYZ(color_matrix, RGB_list[i], tone_curve, XYZ_0[1], XYZ_W[1]))
        
    return np.array(XYZ_list) 

    
def RGB_to_XYZ(Matrix, RGB, tone_curve, Y_min, Y_max, add_black = False):
    rgb = []
    RGB = np.array(RGB)
    RGB[np.isnan(RGB)] = 0
    for i in range(0, len(RGB)):
        if RGB[i] == 0:
            rgb.append(0)
        else:
             rgb.append(RGB[i] / RGB_max)
    
    if tone_curve == "DICOMGSDF":
        trc_index = Cal_dicom_trc_index(Y_min, Y_max)                                  
        add_black = False
        v = np.array(
            [trc_index[int(RGB[0])], 
             trc_index[int(RGB[1])],
             trc_index[int(RGB[2])]])
    elif tone_curve == "PQ":
        trc_index = Cal_HDR_trc_index(tone_curve , Y_min, Y_max)                     
        add_black = False
        v = np.array(
            [trc_index[int(RGB[0])], 
             trc_index[int(RGB[1])],
             trc_index[int(RGB[2])]])
    elif tone_curve == "HLG":    
        trc_index = Cal_HDR_trc_index(tone_curve , Y_min, Y_max)                            
        add_black = False
        v = np.array(
            [trc_index[int(RGB[0])], 
             trc_index[int(RGB[1])],
             trc_index[int(RGB[2])]])
    else:
        tone_curve_float = float(tone_curve) 
        add_black = True
        v = np.array(
            [pow(rgb[0], tone_curve_float), 
             pow(rgb[1], tone_curve_float),
             pow(rgb[2], tone_curve_float)])
    
    v[np.isnan(v)] = 0

    if add_black:
        lv_scale = Y_max / (Matrix.dot([1,1,1]) *  (Y_max - Y_min) + Y_min)[1]
        XYZ = (Matrix.dot(v) * (Y_max - Y_min)) * lv_scale + Y_min
    else:
        lv_scale = Y_max / (Matrix.dot([1,1,1]) * Y_max)[1]
        XYZ = Matrix.dot(v) * Y_max * lv_scale
    return XYZ 


def Cal_HDR_trc_index(hdr_type, lv_min, lv_max):
    trc_index = []
    if hdr_type == "PQ":
        for i in range(0, 1024):
            trc_index.append(PQ_EOTF(i, lv_max))
    elif hdr_type == "HLG" :
        for i in range(0, 1024):
            trc_index.append(HLG_EOTF(i, lv_max, lv_min))
    else:
        return 0
    
    return trc_index
    
def PQ_EOTF(E, lv_max, normailize = True):
    m1 = 0.1593017578125 
    m2 = 78.84375 
    c2 = 18.8515625
    c3 = 18.6875 
    c1 = c3 - c2 + 1
    E /= RGB_max
    
    numerator = max(pow(E, 1.0 / m2) - c1, 0)
    denominator = c2 - c3 * pow(E, 1.0 / m2)
    
    Y = min(10000 * pow(numerator / denominator, 1.0 / m1), lv_max)
    
    if (normailize):
        Y /= lv_max
    
    return Y


def HLG_EOTF(E, lv_max, lv_min, normailize = True):
    E /= RGB_max
    Lw = lv_max
    Lb = lv_min
    
    sys_gamma = 1.2 + 0.42 * log10(𝐿w / 1000)
    
    B = sqrt(3 * pow((Lb/Lw), 1 / sys_gamma))
    
    X = max(0, (1-B) * E + B)
    
    a = 0.17883277
    b = 0.28466892
    c = 0.55991073
    
    if X <= 0.5:
        ootf_e = pow(X, 2) / 3.0
    else:
        ootf_e = (exp((X - c) / a) + b) / 12
    
    Y_s = 0.2627 * E + 0.6780 * E + 0.0593 *E
    Y = pow(Y_s, sys_gamma - 1) * ootf_e
    
    return Y 


def f(t):
    t_o = as_float(np.where(
            t > 0.008856,
            t ** (1.0 / 3.0),
            (903.3 * t + 16.0) / 116.0,
    ))
    return t_o


def XYZ_to_Lab(XYZ, XYZ_r):
    #XYZ_w_TARGET = CCT_to_xyz_nor(6504)
    #XYZ = chromatic_adaptation(XYZ, XYZ_r, XYZ_w_TARGET)
    #XYZ_r = chromatic_adaptation(XYZ_r, XYZ_r, XYZ_w_TARGET)
    
    xr = XYZ[0]/XYZ_r[0]
    yr = XYZ[1]/XYZ_r[1]
    zr = XYZ[2]/XYZ_r[2]
    
    fx = f(xr)
    fy = f(yr)
    fz = f(zr)
    
    L = 116.0 * fy - 16.0
    a = 500.0 * (fx - fy)
    b = 200.0 * (fy - fz)
    
    return np.array([L, a, b]) 


def chromatic_adaptation(XYZ, XYZ_W_raw, XYZ_W_target):
    XYZ_nor = XYZ / XYZ_W_raw[1]
    L_A1 = XYZ_W_raw[1] / np.pi
    L_A2 = XYZ_W_raw[1] / np.pi
        
    XYZ_cp = colour.chromatic_adaptation(
        XYZ_nor, XYZ_W_raw, XYZ_W_target, method = 'CMCCAT2000', L_A1 = L_A1, L_A2 = L_A2)
            
    return XYZ_cp * XYZ_W_raw[1]


def XYZ_to_Luv(XYZ, XYZ_r):
    X = XYZ[0]
    Y = XYZ[1]
    Z = XYZ[2]
    
    X_r = XYZ_r[0]
    Y_r = XYZ_r[1]
    Z_r = XYZ_r[2]
    
    if Y/Y_r <= pow(6/29.0, 3): 
        L = pow(29/3.0, 3) * Y/Y_r
    else:
        L = 116 * pow(Y/Y_r, 1/3.0) - 16
        
    u = (13 * L * ((4 * X / (X + 15 * Y + 3 * Z)) - (4 * X_r / (X_r + 15 * Y_r + 3 * Z_r))))
    v = (13 * L * ((9 * Y / (X + 15 * Y + 3 * Z)) - (9 * Y_r / (X_r + 15 * Y_r + 3 * Z_r))))
    
    return np.array([L, u, v]) 

def xy_to_XYZ(xy, Y):
    X = Y*xy[0]/xy[1]
    Z = Y*(1.0-xy[0]-xy[1])/xy[1]
    
    return np.array([X, Y, Z]) 


def L_to_JND(L):
    if L <= 0:
       JND = 0
    else:
        JND = 71.498068 + 94.593053 * log10(L) + 41.912053 * (pow(log10(L), 2)) + 9.8247004 * (pow(log10(L), 3))+ 0.28175407 * (pow(log10(L), 4)) - 1.1878455 * (pow(log10(L), 5))- 0.18014349 * (pow(log10(L), 6)) + 0.14710899 * (pow(log10(L), 7)) - 0.017046845 * (pow(log10(L), 8));
   
    return JND

def JND_to_L(JND):
    j = JND
    d = -1.3011877 + 0.080242636 * log(j) + 0.13646699 * (pow(log(j), 2)) - 0.025468404 * (pow(log(j), 3)) + 0.0013635334 * (pow(log(j), 4));
    m = 1 - 0.025840191 * log(j) - 0.10320229 * (pow(log(j), 2)) + 0.028745620 * (pow(log(j), 3)) - 0.0031978977 * (pow(log(j), 4)) + 0.00012992634 * (pow(log(j), 5));
    log10L = d / m
    L = pow(10, log10L)
    
    return L


def j_p(Jmax, Jmin, n, p):
    jp = Jmin + (p / (pow(2, n) - 1)) * (Jmax - Jmin)
    return jp;


def Cal_dicom_trc_index(min_lv, max_lv):
    dicom_trc_index = []

    J_max = L_to_JND(max_lv)
    J_min = L_to_JND(min_lv)
    J_min = L_to_JND(1.0)
    n = 10
    
    for i in range(0, 1024):
        jp = j_p(J_max, J_min, n, i)
        
        if jp <= 0 :
            L_D = 0
        else:
            L_D = JND_to_L(jp)
            L_D /= max_lv
        dicom_trc_index.append(L_D)
   
    dicom_trc_index[-1] = 1.0

    return dicom_trc_index


def main():
    raw_color_file = './raw_color.csv'
    raw_color = read_csv_data(raw_color_file)
    raw_color_XYZ = np.vstack([raw_color[:, 3:6]])
    XYZ_R = raw_color_XYZ[1055]
    XYZ_G = raw_color_XYZ[1087]
    XYZ_B = raw_color_XYZ[1119]
    XYZ_0 = raw_color_XYZ[0]
    XYZ_W = raw_color_XYZ[1023]
    
    matrix_native_RGB_XYZ = Build_Matrix_RGB_to_XYZ(XYZ_R, XYZ_G, XYZ_B, XYZ_0, XYZ_W)
    temp_ex_file = './template_comprehensive.xlsx'
    temp_ex = openpyxl.load_workbook(temp_ex_file)
    

    color_data_path =  './color_measure_data/'  
    for fs in os.listdir(color_data_path):
        f = os.path.splitext(fs)[0]
        f = f.split("_")

        gamut = f[0]
        cct = f[1]
        tone_curve = f[2]
        
        measure_color = read_csv_data(color_data_path + fs)
        RGB_list = np.vstack([measure_color[:, 0:3]])
        measure_color_XYZ = np.vstack([measure_color[:, 3:6]])
        target_color_XYZ = XYZ_STD(matrix_native_RGB_XYZ, gamut, cct, tone_curve, measure_color_XYZ[0],  measure_color_XYZ[1023], RGB_list)
        ref_white = target_color_XYZ[1023]

        CCT_list_raw = []
        CCT_list_ca = []
        CCT_list_std = []
        uv_list_raw = []
        uv_list_ca = []
        uv_list_std = []
        delta_uv_list_raw = []
        delta_uv_list_ca = []
        delta_E_list_raw = []
        delta_E_list_ca = []
        JND_list_raw = []
        JND_list_ca = []
        JND_list_std = []
        dumm = 0.0
        
        for i in range(0, len(measure_color_XYZ)):
            XYZ_raw = raw_color_XYZ[i]
            XYZ_ca = measure_color_XYZ[i]
            XYZ_std = target_color_XYZ[i]

            xy_raw = [XYZ_raw[0] / sum(XYZ_raw), XYZ_raw[1] / sum(XYZ_raw)]
            xy_ca = [XYZ_ca[0] / sum(XYZ_ca), XYZ_ca[1] / sum(XYZ_ca)]
            xy_std = [XYZ_std[0] / sum(XYZ_std), XYZ_std[1] / sum(XYZ_std)]
            CCT_list_raw.append(colour.xy_to_CCT(xy_raw))
            CCT_list_ca.append(colour.xy_to_CCT(xy_ca))
            CCT_list_std.append(colour.xy_to_CCT(xy_std))

            Luv_raw = XYZ_to_Luv(XYZ_raw, ref_white)
            Luv_ca = XYZ_to_Luv(XYZ_ca, ref_white)
            Luv_std = XYZ_to_Luv(XYZ_std, ref_white)
            uv_list_raw.append([Luv_raw[1], Luv_raw[2]])
            uv_list_ca.append([Luv_ca[1], Luv_ca[2]])
            uv_list_std.append([Luv_std[1], Luv_std[2]])
            delta_uv_list_raw.append(sqrt(pow(Luv_raw[1] - Luv_std[1],2) + pow(Luv_raw[2] - Luv_std[2],2)))
            delta_uv_list_ca.append(sqrt(pow(Luv_ca[1] - Luv_std[1],2) + pow(Luv_ca[2] - Luv_std[2],2)))
            
            
            Lab_raw = XYZ_to_Lab(XYZ_raw, ref_white)
            Lab_ca = XYZ_to_Lab(XYZ_ca, ref_white)
            Lab_std = XYZ_to_Lab(XYZ_std, ref_white)     
            delta_E_list_raw.append(colour.delta_E_CIE2000(Lab_raw, Lab_std))
            delta_E_list_ca.append(colour.delta_E_CIE2000(Lab_ca, Lab_std))
            
            if i < 1024:
                JND_list_raw.append(L_to_JND(raw_color_XYZ[i][1]))
                JND_list_ca.append(L_to_JND(measure_color_XYZ[i][1]))
                JND_list_std.append(L_to_JND(target_color_XYZ[i][1]))
            
            dumm = dumm + colour.delta_E_CIE2000(Lab_ca, Lab_std)
            
        print("%s_%s_%s - mean dE00:%.5f, max dE00:%.5f" % (gamut, cct, tone_curve, sum(delta_E_list_ca) / len(delta_E_list_ca), max(delta_E_list_ca)))
        
        exs = temp_ex.get_sheet_by_name("data")                
        write_color_data(raw_color_XYZ, 0, 5, 1, exs)
        write_color_data(target_color_XYZ, 0, 27, 1, exs)
        write_color_data(measure_color_XYZ, 0, 44, 1, exs)
        write_color_data(CCT_list_raw, 0, 18, 2, exs)
        write_color_data(CCT_list_std, 0, 37, 2, exs)
        write_color_data(CCT_list_ca, 0, 57, 2, exs)
        write_color_data(uv_list_raw, 0, 15, 1, exs)
        write_color_data(uv_list_ca, 0, 54, 1, exs)
        write_color_data(uv_list_std, 0, 35, 1, exs)
        write_color_data(delta_uv_list_raw, 0, 17, 2, exs)
        write_color_data(delta_uv_list_ca, 0, 56, 2, exs)
        write_color_data(delta_E_list_raw, 0, 19, 2, exs)
        write_color_data(delta_E_list_ca, 0, 58, 2, exs)
        write_color_data(JND_list_raw, 0, 20, 2, exs)
        write_color_data(JND_list_std, 0, 38, 2, exs)
        write_color_data(JND_list_ca, 0, 59, 2, exs)
        exs.cell(1, 65).value = tone_curve
        save_path = './report_data/' + gamut + "_" + cct + '_' + tone_curve + '.xlsx'
        temp_ex.save(save_path)
    
    print("\nFinish!")
    os.system("pause")
main()