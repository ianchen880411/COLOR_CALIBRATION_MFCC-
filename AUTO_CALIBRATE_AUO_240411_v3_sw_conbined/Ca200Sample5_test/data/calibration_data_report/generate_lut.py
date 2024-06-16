# -*- coding: utf-8 -*-
"""
Created on Tue Mar  5 08:53:00 2024

@author: ian88
"""

import os
import numpy as np
import pandas as pd
import openpyxl


M_sRGB_RGB_XYZ = np.array(
            [[0.4124, 0.3576, 0.1805],
             [0.2126, 0.7152, 0.0722],
             [0.0193, 0.1192, 0.9505]])
    
M_ADOBE_RGB_XYZ = np.array(
            [[0.5767309, 0.1855540, 0.1881852],
             [0.2973769, 0.6273491, 0.0752741],
             [0.0270343, 0.0706872, 0.9911085]])
        
M_BT2020_RGB_XYZ = np.array(
            [[0.6370, 0.1446, 0.1689],
             [0.2627, 0.6780, 0.0593],
             [0.0, 0.0281, 1.0610]])

gamut_matrix_list = {'sRGB': M_sRGB_RGB_XYZ, 'ADOBE': M_ADOBE_RGB_XYZ, 'BT2020': M_BT2020_RGB_XYZ}

ebu_test_color_Yuv_list = np.array(
    [[9.60, 0.2526, 0.5013], [37.7, 0.2365, 0.4931], [29.8, 0.2365, 0.4850],
     [29.9, 0.1805, 0.5453], [29.8, 0.1630, 0.4552], [30.1, 0.2088, 0.4155],
     [13.4, 0.1816, 0.5205], [19.3, 0.3246, 0.4972], [43.6, 0.1504, 0.5329],
     [17.2, 0.1789, 0.3707], [6.50, 0.3047, 0.4898], [20.0, 0.1461, 0.5320],
     [6.00, 0.1828, 0.3420], [43.4, 0.2724, 0.5274], [20.0, 0.2348, 0.4031]])

def write_color_data(data, r,c, mode, exs = None):
    if (mode == 1):
        for d in range(0, len(data)):
            for i in range(len(data[d])):
                exs.cell(r+d+2, i+c).value=np.round(data[d][i], 4)
    elif (mode == 2):
        for d in range(0,len(data)):
            exs.cell(d+2, c).value=np.round(data[d], 4)
   

def generate_ebu_test_lut(gamut, LUT, Yuv_list, trc, Y_max):
    lut = []
    gamut_matrix = gamut_matrix_list[gamut]
    
    for i in range(0, len(Yuv_list)):
        Yuv = Yuv_list[i]
        XYZ = Yuv_to_XYZ(Yuv, Y_max)
        RGB = XYZ_to_RGB(gamut_matrix, XYZ, trc, Y_max)
        lut.append(RGB)
        
    return lut


def XYZ_to_RGB(Matrix, XYZ, trc, Y_max):
    XYZ = np.asarray(XYZ)
    XYZ /= Y_max
    m =  np.linalg.inv(Matrix)
    rgb = m.dot(XYZ)
    
    for i in range(0, len(rgb)):
        if rgb[i] < 0:
            rgb[i] = 0
        
    R = np.round(pow((rgb[0]), 1/trc) * 1023, 0)
    G = np.round(pow((rgb[1]), 1/trc) * 1023, 0)
    B = np.round(pow((rgb[2]), 1/trc) * 1023, 0)
    
    return [R, G, B]


def Yuv_to_XYZ(Yuv, Y_max):
    x = 9 * Yuv[1] / (6 * Yuv[1] - 16 * Yuv[2] + 12)
    y = 4 * Yuv[2] / (6 * Yuv[1] - 16 * Yuv[2] + 12)
    Y = Y_max * Yuv[0] / 100.0
    X = Y * x / y
    Z = Y * (1 - x - y )/ y
    XYZ_list = [X, Y, Z]
     
    return XYZ_list


def generate_lut():
    lut = []
    lut_RGB = []
    lut_grayscale = []
    lut_r = []
    lut_g = []
    lut_b = []
    
    for i in range(0, 1024):
        lut_grayscale.append([i, i, i])

    for i in range(32, 1025, 32):
        v = min(i, 1023)
        lut_r.append([v, 0, 0])
        lut_g.append([0, v, 0])
        lut_b.append([0, 0, v])
        
    lut = np.concatenate((lut_grayscale, lut_r),axis=0)  
    lut = np.concatenate((lut, lut_g),axis=0)  
    lut = np.concatenate((lut, lut_b),axis=0)  
     
    for r in range(0, 1025, 128):
        for g in range(0, 1025, 128):
            for b in range(0, 1025, 128):
                if r == g and g == b:
                    pass
                elif (r + g + b) / 3.0 == max(max(r,g),b) / 3.0:
                    pass
                else:
                    r = min(r, 1023)
                    g = min(g, 1023)
                    b = min(b, 1023)
                    lut_RGB.append([r, g, b])
    
    lut = np.concatenate((lut, lut_RGB),axis=0)
    
    gamut = "sRGB"
    trc = 2.35
    Y_max = 100
    lut_ebu = generate_ebu_test_lut(gamut, lut, ebu_test_color_Yuv_list, trc, Y_max)
    lut = np.concatenate((lut, lut_ebu),axis=0)
    
    return lut

lut = generate_lut()

temp_ex_file = './test_lut.xlsx'
temp_ex = openpyxl.load_workbook(temp_ex_file)
exs = temp_ex.get_sheet_by_name("data")     
write_color_data(lut, 0, 1, 1, exs)
temp_ex.save(temp_ex_file)